"""
EXCEL SETUP - Creates Excel file based on stocks_config.json
FIXED: Sector as a column (like Day), not in header!
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta, time as dt_time
import json
import os

STOCKS_CONFIG_FILE = 'stocks_config.json'
EXCEL_FILE = 'Stock_Tracker_Fixed.xlsx'

TIME_SLOTS = [
    dt_time(9, 0), dt_time(9, 15), dt_time(9, 30), dt_time(9, 45),
    dt_time(10, 0), dt_time(10, 15), dt_time(10, 30), dt_time(10, 45),
    dt_time(11, 0), dt_time(11, 15), dt_time(11, 30), dt_time(11, 45),
    dt_time(12, 0), dt_time(12, 15), dt_time(12, 30), dt_time(12, 45),
    dt_time(13, 0), dt_time(13, 15), dt_time(13, 30), dt_time(13, 45),
    dt_time(14, 0), dt_time(14, 15), dt_time(14, 30), dt_time(14, 45),
    dt_time(15, 0), dt_time(15, 15), dt_time(15, 30)
]


def load_stocks_config():
    """Load stocks from JSON configuration file"""
    if not os.path.exists(STOCKS_CONFIG_FILE):
        print(f"ERROR: Configuration file '{STOCKS_CONFIG_FILE}' not found!")
        print("Please run stock_manager.py to add stocks first.")
        return {}
    
    with open(STOCKS_CONFIG_FILE, 'r') as f:
        config = json.load(f)
    
    stocks = {item['symbol']: item['sector'] for item in config['stocks']}
    return stocks


def create_excel_structure(start_date_str='01-01-2026', num_days=30):
    """
    Create Excel file with column-based structure
    Layout: Date | Time | Day | Sector | STOCK1 | STOCK2 | ...
    """
    
    print("=" * 70)
    print("EXCEL SETUP - SECTOR AS COLUMN (FIXED!)")
    print("=" * 70)
    print()
    
    # Load stocks
    STOCKS = load_stocks_config()
    
    if not STOCKS:
        print("No stocks loaded. Exiting.")
        return
    
    print(f"✓ Loaded {len(STOCKS)} stocks")
    print()
    
    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Stock Data"
    
    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Headers
    sheet['A1'] = 'Date'
    sheet['B1'] = 'Time'
    sheet['C1'] = 'Day'
    sheet['D1'] = 'Sector'  # NEW: Sector column header
    
    # Stock name headers (starting from column E)
    col = 5
    for symbol in STOCKS.keys():
        stock_name = symbol.replace('.NS', '').replace('.BO', '')
        sheet.cell(row=1, column=col, value=stock_name)
        col += 1
    
    # Format header row
    for col in range(1, 5 + len(STOCKS)):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Generate date rows
    start_date = datetime.strptime(start_date_str, '%d-%m-%Y').date()
    row = 2
    
    for day_offset in range(num_days):
        current_date = start_date + timedelta(days=day_offset)
        
        # Skip weekends
        if current_date.weekday() >= 5:
            continue
        
        date_str = current_date.strftime('%d-%m-%Y')
        day_name = current_date.strftime('%A')
        
        # For each time slot
        for time_slot in TIME_SLOTS:
            time_str = time_slot.strftime('%H:%M')
            
            # Fill basic columns
            sheet.cell(row=row, column=1, value=date_str)
            sheet.cell(row=row, column=2, value=time_str)
            sheet.cell(row=row, column=3, value=day_name)
            sheet.cell(row=row, column=4, value='')  # Sector filled by populate script
            
            # Format basic columns
            for col in [1, 2, 3, 4]:
                cell = sheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                cell.font = Font(name='Arial', size=10)
            
            # Add borders to stock columns
            for col in range(5, 5 + len(STOCKS)):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
            
            row += 1
    
    # Column widths
    sheet.column_dimensions['A'].width = 12  # Date
    sheet.column_dimensions['B'].width = 8   # Time
    sheet.column_dimensions['C'].width = 12  # Day
    sheet.column_dimensions['D'].width = 18  # Sector
    
    # Stock columns
    for col in range(5, 5 + len(STOCKS)):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 12
    
    # Freeze panes (after Sector column)
    sheet.freeze_panes = 'E2'
    
    # Save
    wb.save(EXCEL_FILE)
    
    print(f"✓ Excel file created: {EXCEL_FILE}")
    print(f"✓ Layout: Date | Time | Day | Sector | {len(STOCKS)} stock columns")
    print(f"✓ Total rows: {row - 1}")
    print()
    print("=" * 70)
    print("✓ SETUP COMPLETE!")
    print("=" * 70)
    print()
    print("Next steps:")
    print("  1. Review the Excel file")
    print("  2. Run populate_modular.py to fetch data")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        start_date = sys.argv[1]
    else:
        start_date = '01-01-2026'
    
    if len(sys.argv) > 2:
        num_days = int(sys.argv[2])
    else:
        num_days = 30
    
    create_excel_structure(start_date, num_days)
