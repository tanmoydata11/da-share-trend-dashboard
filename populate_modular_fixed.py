"""
MODULAR Stock Data Fetcher - SECTOR COLUMN FIXED!
- Sector shown in each row (like Day column)
- Stocks loaded from stocks_config.json
- Fast parallel fetching
- Professional column-based layout
"""
import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta, time as dt_time
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import json
import os

# ============ CONFIGURATION FILE ============
STOCKS_CONFIG_FILE = 'stocks_config.json'
EXCEL_FILE = 'Stock_Tracker_Fixed.xlsx'

# ============ TIME SLOTS ============
TIME_SLOTS = [
    dt_time(9, 0), dt_time(9, 15), dt_time(9, 30), dt_time(9, 45),
    dt_time(10, 0), dt_time(10, 15), dt_time(10, 30), dt_time(10, 45),
    dt_time(11, 0), dt_time(11, 15), dt_time(11, 30), dt_time(11, 45),
    dt_time(12, 0), dt_time(12, 15), dt_time(12, 30), dt_time(12, 45),
    dt_time(13, 0), dt_time(13, 15), dt_time(13, 30), dt_time(13, 45),
    dt_time(14, 0), dt_time(14, 15), dt_time(14, 30), dt_time(14, 45),
    dt_time(15, 0), dt_time(15, 15), dt_time(15, 30)
]

# ============ COLORS ============
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
NEUTRAL_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')


def load_stocks_config():
    """Load stocks from JSON configuration file"""
    if not os.path.exists(STOCKS_CONFIG_FILE):
        print(f"ERROR: Configuration file '{STOCKS_CONFIG_FILE}' not found!")
        return {}
    
    try:
        with open(STOCKS_CONFIG_FILE, 'r') as f:
            config = json.load(f)
        
        stocks = {item['symbol']: item['sector'] for item in config['stocks']}
        print(f"✓ Loaded {len(stocks)} stocks from {STOCKS_CONFIG_FILE}")
        return stocks
        
    except Exception as e:
        print(f"ERROR: Failed to load stocks: {e}")
        return {}


def get_intraday_prices(symbol, target_date):
    """Fetch intraday prices for one stock on one date"""
    try:
        ticker = yf.Ticker(symbol)
        start = target_date
        end = target_date + timedelta(days=1)
        data = ticker.history(start=start, end=end, interval='1m')
        
        if data.empty:
            return {}
        
        try:
            data.index = data.index.tz_convert('Asia/Kolkata')
        except:
            pass
        
        data = data[data.index.date == target_date]
        if data.empty:
            return {}
        
        prices = {}
        for slot in TIME_SLOTS:
            start_time = (datetime.combine(target_date, slot) - timedelta(minutes=7, seconds=30)).time()
            end_time = (datetime.combine(target_date, slot) + timedelta(minutes=7, seconds=30)).time()
            mask = (data.index.time >= start_time) & (data.index.time <= end_time)
            slot_data = data[mask]
            
            if not slot_data.empty:
                prices[slot] = round(slot_data['Close'].iloc[-1], 2)
        
        return prices
    except:
        return {}


def fetch_stock_day(symbol, date):
    """Fetch one stock for one day"""
    prices = get_intraday_prices(symbol, date)
    return (symbol, date, prices)


def populate_fixed(start_date_str='01-01-2026', end_date_str=None, max_workers=5):
    """
    Populate Excel with stock data
    FIXED: Fills Sector column for each row
    """
    
    print("=" * 70)
    print("MODULAR STOCK TRACKER - SECTOR COLUMN FIXED!")
    print("=" * 70)
    print()
    
    # Load stocks
    STOCKS = load_stocks_config()
    if not STOCKS:
        print("No stocks loaded. Exiting.")
        return
    
    start_date = datetime.strptime(start_date_str, '%d-%m-%Y').date()
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%d-%m-%Y').date()
    else:
        end_date = datetime.now().date()
    
    # Get trading days
    trading_days = []
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:
            trading_days.append(current)
        current += timedelta(days=1)
    
    print(f"Date Range: {start_date} to {end_date}")
    print(f"Trading Days: {len(trading_days)}")
    print(f"Stocks: {len(STOCKS)}")
    print(f"Parallel Workers: {max_workers}")
    print(f"Total tasks: {len(trading_days) * len(STOCKS)}")
    print()
    
    # Load Excel
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Excel file '{EXCEL_FILE}' not found!")
        print("Please run setup_excel_fixed.py first!")
        return
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    # Get stock column mapping (now starting from column 5 = E)
    stock_columns = {}
    for col in range(5, 200):  # Starting from column E (5)
        stock_name = sheet.cell(row=1, column=col).value
        if not stock_name:
            break
        # Find full symbol
        for symbol in STOCKS.keys():
            if symbol.replace('.NS', '').replace('.BO', '') == stock_name:
                stock_columns[symbol] = col
                break
    
    print(f"Found {len(stock_columns)} stock columns in Excel")
    print()
    
    # Warn about missing stocks
    missing_in_excel = set(STOCKS.keys()) - set(stock_columns.keys())
    if missing_in_excel:
        print("⚠ WARNING: These stocks are in config but NOT in Excel:")
        for s in missing_in_excel:
            print(f"  - {s.replace('.NS', '').replace('.BO', '')}")
        print("  Please run setup_excel_fixed.py to add them.")
        print()
    
    print("⚡ Starting parallel fetch...")
    print()
    
    # Create tasks
    tasks = []
    for date in trading_days:
        for symbol in stock_columns.keys():
            tasks.append((symbol, date))
    
    # Fetch all data
    all_data = {}
    completed = 0
    total = len(tasks)
    start_time = time.time()
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_task = {
            executor.submit(fetch_stock_day, symbol, date): (symbol, date)
            for symbol, date in tasks
        }
        
        for future in as_completed(future_to_task):
            symbol, date, prices = future.result()
            
            if date not in all_data:
                all_data[date] = {}
            all_data[date][symbol] = prices
            
            completed += 1
            if completed % 10 == 0 or completed == total:
                elapsed = time.time() - start_time
                rate = completed / elapsed if elapsed > 0 else 0
                remaining = (total - completed) / rate if rate > 0 else 0
                print(f"Progress: {completed}/{total} ({completed/total*100:.1f}%) | "
                      f"Rate: {rate:.1f}/sec | ETA: {remaining:.0f}s")
    
    print()
    print("✓ All data fetched!")
    print(f"⏱️  Fetch time: {time.time() - start_time:.1f}s")
    print()
    print("Writing to Excel with colors and sectors...")
    
    write_start = time.time()
    
    # Write data with color coding
    for date in sorted(all_data.keys()):
        date_str = date.strftime('%d-%m-%Y')
        
        # Find starting row for this date
        start_row = None
        for row in range(2, 10000):  # Starting from row 2 (headers in row 1)
            if sheet.cell(row=row, column=1).value == date_str:
                start_row = row
                break
        
        if not start_row:
            continue
        
        # For each time slot in this date
        for time_idx, time_slot in enumerate(TIME_SLOTS):
            current_row = start_row + time_idx
            
            # FILL SECTOR COLUMN (column D = 4)
            # Get sector from the first stock that has data for this row
            sector_filled = False
            for symbol in stock_columns.keys():
                if symbol in all_data[date]:
                    sector = STOCKS[symbol]
                    sheet.cell(row=current_row, column=4, value=sector)
                    sector_filled = True
                    break
            
            # For each stock
            for symbol, prices in all_data[date].items():
                if symbol not in stock_columns:
                    continue
                
                stock_col = stock_columns[symbol]
                price = prices.get(time_slot)
                
                if price:
                    cell = sheet.cell(row=current_row, column=stock_col)
                    cell.value = price
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    
                    # COLOR CODING: Compare with previous time slot
                    if time_idx > 0:
                        prev_row = current_row - 1
                        prev_cell = sheet.cell(row=prev_row, column=stock_col)
                        prev_price = prev_cell.value
                        
                        if prev_price is not None and isinstance(prev_price, (int, float)):
                            if price > prev_price:
                                cell.fill = GREEN_FILL
                            elif price < prev_price:
                                cell.fill = RED_FILL
                            else:
                                cell.fill = NEUTRAL_FILL
                        else:
                            cell.fill = NEUTRAL_FILL
                    else:
                        cell.fill = NEUTRAL_FILL
    
    wb.save(EXCEL_FILE)
    
    print(f"✓ Excel updated!")
    print(f"⏱️  Write time: {time.time() - write_start:.1f}s")
    print(f"⏱️  TOTAL: {time.time() - start_time:.1f}s")
    print()
    print("=" * 70)
    print("✓ COMPLETE!")
    print("=" * 70)


def update_today_fixed():
    """Fast update for today only"""
    today = datetime.now().date()
    
    if today.weekday() >= 5:
        print("Weekend - no market data")
        return
    
    today_str = today.strftime('%d-%m-%Y')
    populate_fixed(today_str, today_str, max_workers=7)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == '--today':
        update_today_fixed()
    else:
        populate_fixed('01-01-2026', '31-12-2026', max_workers=5)  # CHANGED: Full year
